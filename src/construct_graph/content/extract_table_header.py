import logging
import pickle
from typing import Any, Tuple, Optional
import openpyxl
import pandas as pd
import pymysql
# Removed unused import xlrd
import shutil
import os
# Removed unused imports dateutil.parser, date, datetime
import numpy as np
import sys # Added for csv.field_size_limit

# ---------------- Configuration Constants ------------------
# Base directory for data files
DATA_DIR = 'data_search_e_data'
# Temporary directory for intermediate files
TEMP_DIR = '.temp'
# Subdirectory within TEMP_DIR for temporary file storage during processing
TEMP_STORE_DIR = '.temp/temp'
# Output file path for the final results (Pickle file)
OUTPUT_FILE = 'output/table_pattern.pkl'

# Supported file formats for table processing
SUPPORTED_FORMATS = ('csv', 'xls', 'xlsx')

# MySQL Database Configuration
MYSQL_CONFIG = {
    "host": "...",
    "port": 3306,
    "user": "...",
    "password": "...",
    "database": "..."
}

# Log file configuration
LOG_FILE = 'logs/table_pattern.log'
LOG_LEVEL = logging.INFO # Set minimum logging level

# ---------------- Logging Setup ------------------
# Ensure logs directory exists
os.makedirs('logs', exist_ok=True)
# Configure logging to capture errors and info to a file and the console
logging.basicConfig(
    level=LOG_LEVEL,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE), # Log messages to a file
        logging.StreamHandler() # Log messages to the console
    ]
)

# ---------------- Helper Functions ------------------

# Removed is_valid_date function as it was commented out and unused

def count_letters(string: str) -> int:
    """Counts the number of alphabetic characters in a string."""
    count = 0
    for char in string:
        if char.isalpha():
            count += 1
    return count

def is_number(cell_value: Any) -> bool:
    """
    Checks if a cell value represents a number.
    Handles non-string types, short strings, formulas, percentages,
    and uses numpy.isreal for broader numeric check.
    """
    # If not a string, assume it's a number type by pandas/openpyxl
    if not isinstance(cell_value, str):
        return True
    # Short strings (e.g., single digits, symbols) might be numbers
    if len(cell_value) <= 2:
        return True
    # Check for formula or percentage-like strings with few letters
    if cell_value.startswith('=') or ('%' in cell_value and count_letters(cell_value) < 4):
        return True
    # Check if the string represents a digit, decimal, numeric, or real number
    if str(cell_value).isdigit() or str(cell_value).isdecimal() \
            or str(cell_value).isnumeric() or np.isreal(str(cell_value)):
        return True
    # Otherwise, assume it's not purely a number based on these checks
    return False

# ---------------- XLSX Feature Extraction ------------------
def xlsx_features(file_path: str) -> tuple:
    """
    Extracts structural and content features from an XLSX file sheet by sheet.
    Includes detecting if a sheet is a chart sheet, row/column counts,
    content of first column/row, and patterns in row emptiness/numeric content.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        logging.error(f"Error loading XLSX file {file_path}: {e}")
        # Return default empty-like structures on failure
        return [], [], [], [], [], [], [], [], []

    is_sheet = [0] * len(wb.sheetnames) # Flag: 1 if a sheet, 0 if a chartsheet or effectively empty
    sheet_rows = [-1] * len(wb.sheetnames) # Number of rows with content
    table_type = [0] * len(wb.sheetnames) # Flag: 1 if simple header, 0 if complex header or no clear header
    last_bold_row = [-1] * len(wb.sheetnames) # Index of the last row with bold text (0-based within processed rows)
    row_contents = [] # Features derived from row contents (numeric ratio, non-none count) for early rows
    col_a_contents = [] # Content of column 'A'
    col_a_none_counts = [] # Count of None values in column 'A'
    col_contents = [] # Features derived from first contentful column (numeric ratio, none ratio)
    sheetnames = wb.sheetnames # List of sheet names

    for sheet_id, sheet_name in enumerate(sheetnames):
        sheet = wb[sheet_name]

        # Skip chart sheets or sheets with only 1 row
        if str(type(sheet)) == "<class 'openpyxl.chartsheet.chartsheet.Chartsheet'>" or sheet.max_row == 1:
            row_contents.append(None)
            col_a_contents.append(None)
            col_a_none_counts.append(None)
            col_contents.append(None)
            continue # Move to the next sheet

        is_sheet[sheet_id] = 1
        sheet_rows[sheet_id] = sheet.max_row

        # Determine max column with content
        max_col = 0
        # Iterate through columns to find the last one with at least one non-None value
        for i, col in enumerate(sheet.iter_cols()):
            col_value = [cell.value for cell in col]
            if col_value.count(None) != len(col_value):
                max_col = i + 1
            else:
                # If an empty column is encountered, assume it's the effective end
                # This might stop early if there are gaps, but matches original logic
                break

        # Basic header detection: Check if the first row has no None values (simple header assumption)
        # Only check if there's more than one column
        if max_col > 1:
            first_row_values = [cell.value for cell in sheet[1][:max_col]]
            if first_row_values.count(None) == 0:
                table_type[sheet_id] = 1 # Mark as likely simple header

        # Collect non-None counts for the first 25 rows (or fewer if sheet is small)
        all_row_nones = []
        for i in range(1, min(sheet.max_row + 1, 26)):
            all_row_nones.append([cell.value for cell in sheet[i]].count(None))

        # Collect row features (numeric ratio, non-none count) for the first 15 rows
        # Also detect last row with bold text (0-based index within processed rows)
        current_row_features = []
        current_row_none_counts = []
        current_row_lengths = []
        num_processed_rows = 0 # Counter for rows added to current_row_features

        for i in range(1, min(sheet.max_row + 1, 16)): # Process up to 15 rows
            row_values = []
            numeric_count = 0
            row_len = 0 # Count of non-None cells in this specific row
            total_cells_in_row = 0 # Total cells in this row up to sheet.max_column
            try:
                # Iterate through cells in the row
                for j in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=i, column=j)
                    row_values.append(cell.value)
                    total_cells_in_row += 1

                    if cell.value is not None:
                         row_len += 1 # Count non-none cells

                    if cell.value is not None and len(str(cell.value).strip()) > 0 \
                            and (cell.data_type in ['n', 'f', 'b'] or is_number(cell.value)):
                        numeric_count += 1

                    # Check for bold font style
                    if cell.font and cell.font.bold:
                        last_bold_row[sheet_id] = num_processed_rows # Store 0-based index within current_row_features

                num_none_in_row = row_values.count(None)
                # Additional checks for values that might be treated as None/empty in content analysis
                num_none_in_row += row_values.count('...')
                num_none_in_row += row_values.count('. . .')
                num_none_in_row += row_values.count('--')
                num_none_in_row += row_values.count('N/A')

                if row_len > 0: # Only consider rows with some content
                    numeric_ratio = numeric_count / row_len if row_len > 0 else 0 # Ratio of numeric cells among non-none
                    current_row_features.append(row_values + [[num_none_in_row, numeric_ratio]]) # Append features to row values
                    current_row_none_counts.append(num_none_in_row)
                    current_row_lengths.append(total_cells_in_row)
                    num_processed_rows += 1
                else:
                    # Optionally break if several consecutive empty rows are found - not implemented in original, keeping as-is
                    pass

            except Exception as e:
                logging.warning(f"Error processing row {i} in sheet '{sheet_name}' of {file_path}: {e}")
                continue # Continue with the next row

        row_contents.append(current_row_features) # Store features for this sheet's processed rows

        # Get content and none count of column 'A'
        col_a_values = [cell.value for cell in sheet['A']]
        col_a_contents.append(col_a_values)
        col_a_none_counts.append(col_a_values.count(None)) # Store count of None in col A

        # Collect column features (numeric ratio, none ratio) for the first column with content
        current_col_features = None # Initialize to None
        # Check if column 'A' has content (not all None) and is not a Metadata sheet
        if col_a_contents[sheet_id] is not None and sheet_name.lower() not in ['metadata'] \
                and col_a_contents[sheet_id].count(None) != len(col_a_contents[sheet_id]):
            # Iterate through columns to find the first one with content
            for i, column in enumerate(sheet.iter_cols()):
                col_values = []
                col_numeric_count = 0
                col_none_count = 0
                col_total_cells = 0
                try:
                    for cell in column:
                        col_values.append(cell.value)
                        col_total_cells += 1
                        if cell.value is not None:
                            if len(str(cell.value).strip()) > 0 \
                                    and (cell.data_type in ['n', 'f', 'b'] or is_number(cell.value)):
                                col_numeric_count += 1
                        if cell.value is None:
                            col_none_count += 1
                except Exception as e:
                    logging.warning(f"Error processing column {i+1} in sheet '{sheet_name}' of {file_path}: {e}")
                    continue # Continue to next column

                if col_total_cells > 0:
                    col_numeric_ratio = col_numeric_count / col_total_cells if col_total_cells > 0 else 0
                    col_none_ratio = col_none_count / col_total_cells if col_total_cells > 0 else 0
                    # Store features for this column
                    current_col_features = col_values + [[col_none_ratio, col_numeric_ratio]]
                    # Break after the first column with content is processed
                    break
                # If loop finishes without finding a column with content, current_col_features remains None

        col_contents.append(current_col_features) # Store features for the first contentful column


        # Refine table type based on row content patterns (e.g., consistent non-none counts)
        # This logic attempts to detect if the top rows form a consistent header block
        # by analyzing `current_row_none_counts` derived from the first 15 rows.
        if current_row_none_counts: # Check if any row features were collected
            first_row_none_count = current_row_none_counts[0]
            # Assuming first_row_len in original logic meant total cells in the first row
            first_row_total_cells = current_row_lengths[0] if current_row_lengths else 0

            # Conditions suggesting a simple header type (fixed the parenthesis error here)
            if ((current_row_none_counts[:3].count(first_row_none_count) == len(current_row_none_counts[:3]) # First 3 rows have same none count
                 and first_row_none_count != first_row_total_cells - 1) # and first row is not almost empty (only 1 non-none cell)
                or current_row_none_counts.count(first_row_none_count) == len(current_row_none_counts) # All processed rows have same none count
                or first_row_none_count == 0 # First row has no empty cells
                or (first_row_none_count != first_row_total_cells - 1 # first row is not almost empty
                    and all_row_nones and first_row_none_count == sorted(all_row_nones, key=lambda item: (all_row_nones.count(item), item))[-1])):
                # AND its none count is the most frequent among first 25 rows
                table_type[sheet_id] = 1 # Mark as simple header

    # Return collected features and sheet names
    # col_2_none from original was duplicate of col_a_none_counts, so returning col_a_none_counts
    return is_sheet, table_type, sheet_rows, row_contents, col_a_contents, last_bold_row, col_a_none_counts, col_contents, sheetnames


# ---------------- Table Feature and Header Extraction ------------------
def table_features(file_path: str, file_format: str) -> tuple:
    """
    Extracts features from a table file (xls, xlsx).
    Converts .xls to .xlsx if necessary using libreoffice.
    Calls xlsx_features for the main extraction logic.
    Handles temporary file creation and cleanup.
    """
    tmpfile_store_path = TEMP_STORE_DIR
    # Ensure temporary storage directory exists
    os.makedirs(tmpfile_store_path, exist_ok=True)

    processing_file_path = file_path # Path to the file currently being processed (might be temp .xlsx)
    temp_file_created = False # Flag to indicate if a temporary file needs cleanup

    if file_format == 'xls':
        # Get the base filename without extension
        file_name = os.path.basename(file_path)
        base_name = os.path.splitext(file_name)[0]
        # Define the path for the temporary .xlsx file
        tmp_file = os.path.join(tmpfile_store_path, base_name + '.xlsx')
        
        logging.info(f"Converting XLS file {file_path} to XLSX for processing...")
        try:
            # Use libreoffice to convert .xls to .xlsx
            # Assumes libreoffice is installed and in the system's PATH
            # --headless runs without GUI, --convert-to specifies output format
            # --outdir specifies output directory
            result = os.system(f"libreoffice --headless --convert-to xlsx \"{file_path}\" --outdir \"{tmpfile_store_path}\"")
            if result != 0:
                 logging.error(f"LibreOffice conversion failed for {file_path}. Return code: {result}")
                 # Attempt to load with pandas as fallback? Original code didn't, stick to original logic.
                 # Return empty features on conversion failure
                 return [], [], [], [], [], [], [], [], []

            processing_file_path = tmp_file # Now process the temporary XLSX file
            temp_file_created = True # Mark for cleanup

            # Add a small delay to ensure the file is written completely
            # import time
            # time.sleep(1) # Optional delay

        except Exception as e:
            logging.error(f"Error during LibreOffice conversion for {file_path}: {e}")
            # Return empty features on failure
            return [], [], [], [], [], [], [], [], []

    elif file_format == 'xlsx':
        # If already XLSX, process directly
        processing_file_path = file_path
        temp_file_created = False # No temp file created

    else:
        # Should not happen given SUPPORTED_FORMATS check in calling function, but defensive
        logging.error(f"Unsupported file format '{file_format}' for feature extraction: {file_path}")
        return [], [], [], [], [], [], [], [], []

    features = [], [], [], [], [], [], [], [], [] # Initialize features structure
    try:
        # Call the XLSX specific feature extraction function
        features = xlsx_features(processing_file_path)
    except Exception as e:
        logging.error(f"Error extracting features from file {processing_file_path}: {e}")
        # Features structure will be the initialized empty one

    finally:
        # Clean up the temporary XLSX file if it was created
        if temp_file_created and os.path.exists(processing_file_path):
            try:
                os.remove(processing_file_path)
                logging.debug(f"Cleaned up temporary file: {processing_file_path}")
            except Exception as e:
                logging.error(f"Error cleaning up temporary file {processing_file_path}: {e}")

    return features


def extract_header(file_path: str, file_format: str, tmp_dir: str) -> tuple:
    """
    Extracts header information from a table file by analyzing its structure
    and content features obtained from table_features.
    Handles temporary file copying for processing.
    """
    copied_file_path = None # Path to the file copy in the temporary directory
    temp_copy_created = False # Flag to indicate if a temporary copy needs cleanup

    try:
        # Determine the extension for the temporary copy
        tmp_extension = '.xlsx' if file_format == 'xls' else os.path.splitext(file_path)[1]
        # Create a temporary copy in the specified temp directory
        copied_file_name = os.path.basename(file_path) + tmp_extension # Add .xlsx extension for xls copies
        copied_file_path = os.path.join(tmp_dir, copied_file_name)

        logging.debug(f"Copying file {file_path} to temporary location {copied_file_path} for header extraction.")
        shutil.copy(file_path, copied_file_path)
        temp_copy_created = True # Mark for cleanup

        # Extract features from the copied file
        features = table_features(copied_file_path, file_format)
        # Unpack the returned features
        is_sheet, table_type, sheet_rows, row_contents, col_a_contents, last_bold_row, col_a_none_counts, col_contents, sheetnames = features

        sheet_num = len(is_sheet) # Total number of sheets processed

        all_sheet_row_headers = [] # List to store extracted row headers per sheet
        all_sheet_col_headers = [] # List to store extracted column headers per sheet

        # Iterate through each sheet's features
        for i in range(sheet_num):
            # If the sheet was skipped during feature extraction, append None
            if is_sheet[i] == 0:
                all_sheet_row_headers.append(None)
                all_sheet_col_headers.append(None)
                continue # Move to the next sheet

            # Initialize header lists for the current sheet
            current_sheet_row_headers = []
            current_sheet_col_headers = []

            # --- Logic to Determine Row Headers ---
            if table_type[i] == 1: # Simple header type (likely CSV-like)
                # If row_contents exist and the first row has features, take the first row as header
                if row_contents and row_contents[i] and row_contents[i][0]:
                    # Exclude the feature list appended at the end of the row values
                    header_row_values = row_contents[i][0][:-1]
                    current_sheet_row_headers.append([item for item in header_row_values if item is not None]) # Filter None values

            else: # Complex header type
                # This logic attempts to find the boundary of the header rows
                # by analyzing patterns in row content (e.g., emptiness, numeric ratio)
                last_header_row_idx = 0 # Index of the last row considered part of the header
                start_row_for_analysis = 0 # Starting row index for pattern analysis

                if row_contents and row_contents[i]: # Check if row features exist for this sheet
                     # Find the first row with many empty cells (potential start of data after header)
                    for j in range(len(row_contents[i])):
                        # Check if the empty cell count ratio is high (e.g., row is mostly empty)
                        # The check [j][-1][0] == len([j])-2 means empty count == total - 2 (allowing 2 non-empty, one for value, one for features)
                        # This condition seems intended to find initial empty/metadata rows
                        if row_contents[i][j][-1][0] == len(row_contents[i][j]) - 2:
                             start_row_for_analysis = j + 1 # Start analysis from the next row
                        else:
                             break # Break on the first row that is not mostly empty

                    # Find the last row of the header block based on consistent empty cell counts in subsequent rows
                    for j in range(start_row_for_analysis, len(row_contents[i]) - 2):
                        # Look for 3 consecutive rows with the same empty cell count
                        # and where that count is not too low (suggesting data rows)
                        if (row_contents[i][j][-1][0] == row_contents[i][j + 1][-1][0] == row_contents[i][j + 2][-1][0]
                            and row_contents[i][j][-1][0] <= 0.5 * len(row_contents[i][j][:-1])): # Empty count is less than half the columns
                            last_header_row_idx = j # This row is the last of the header block
                            break # Found the header boundary

                    # Alternative/supplementary logic to find a potential header boundary
                    min_none_num, min_none_idx = 10000, 500 # Initialize with high values
                    # Find the row with the minimum non-empty count among processed rows
                    # and where the numeric ratio is low (< 0.5, suggesting text headers)
                    for j in range(len(row_contents[i])):
                        none_num = row_contents[i][j][-1][0] # Number of None values in the row
                        numeric_ratio = row_contents[i][j][-1][1] # Ratio of numeric cells
                        if none_num < min_none_num and numeric_ratio < 0.5:
                            min_none_num = none_num
                            min_none_idx = j # Store the index of this row

                    # Use the row with the minimum non-empty count as a potential header boundary index
                    # if it's lower than the boundary found by the consecutive empty rows logic
                    idx = min_none_idx

                    # Consider the last bold row as a potential header boundary if other methods didn't find one
                    # or if the bold row is later than the found boundaries
                    if (last_header_row_idx == 0 or idx == 0) and last_bold_row[i] != -1:
                        # Use the last bold row index (0-based) if no other clear header boundary found
                        # Convert last_bold_row which is num_processed_rows to an index in row_contents
                        # Assuming last_bold_row corresponds to the index within row_contents
                        if last_bold_row[i] < len(row_contents[i]):
                             idx = last_bold_row_idx = last_bold_row[i]
                        else:
                             logging.warning(f"Last bold row index {last_bold_row[i]} out of bounds for row_contents length {len(row_contents[i])} in sheet {i}.")
                             last_bold_row_idx = 0 # Default to 0 if out of bounds


                    # If the consecutive empty row logic didn't find a boundary, use the min_none_idx
                    if last_header_row_idx == 0:
                        last_header_row_idx = idx # Use min_none_idx as the header boundary

                    # Collect all rows from the beginning up to the determined header boundary index
                    # where the numeric ratio is less than 0.5 (indicating non-numeric header-like content)
                    # Take the minimum of (min_none_idx + 1, last_header_row_idx + 1) and the actual number of processed rows
                    # This defines the range of rows considered as headers
                    header_rows_end_index = min(min(idx + 1, last_header_row_idx + 1), len(row_contents[i]))

                    for j in range(header_rows_end_index):
                         # Append row values to header list if numeric ratio is low
                         if row_contents[i][j][-1][1] < 0.5:
                              header_row_values = row_contents[i][j][:-1]
                              current_sheet_row_headers.append([item for item in header_row_values if item is not None]) # Filter None

                         elif row_contents[i][j][-1][1] > 0.7:
                              # If a row with high numeric content (>0.7) is encountered, assume header ends before this row
                              logging.debug(f"Stopping row header collection at row {j} due to high numeric content.")
                              break # Stop collecting rows if high numeric content is found

            # --- Logic to Determine Column Headers ---
            # Check if column features exist for this sheet, and the first column has features,
            # and its numeric ratio is low (< 0.1, suggesting non-numeric column headers)
            # and the column length is not excessively large (< 70, a heuristic to avoid treating data columns as headers)
            if col_contents and col_contents[i] is not None and col_contents[i]:
                col_numeric_ratio = col_contents[i][-1][1]
                col_total_cells = len(col_contents[i]) - 1 # Subtract feature list length
                # Check if numeric ratio is low and column is not excessively long
                if col_numeric_ratio < 0.1 and col_total_cells < 70:
                     # Take the content of the first column features, excluding the feature list
                     col_list = [item for item in col_contents[i][:-1] if item is not None] # Filter None
                     if col_list: # Only append if the list is not empty after filtering
                         current_sheet_col_headers.append(col_list)
                     else:
                         # Append None if the filtered list is empty
                         current_sheet_col_headers.append(None)
                else:
                     # Append None if the column doesn't meet criteria for being a header
                     current_sheet_col_headers.append(None)
            else:
                 # Append None if col_contents don't exist or first column has no features
                 current_sheet_col_headers.append(None)


            # Append the extracted headers for the current sheet to the overall lists
            all_sheet_row_headers.append(current_sheet_row_headers)
            all_sheet_col_headers.append(current_sheet_col_headers)

    except Exception as e:
        logging.error(f"Error extracting header features from file {file_path}: {e}", exc_info=True)
        # Return empty header structures on error
        return [], [], [] # table_type, all_sheet_row_headers, all_sheet_col_headers

    finally:
        # Clean up the temporary file copy if it was created
        if temp_copy_created and copied_file_path and os.path.exists(copied_file_path):
            try:
                os.remove(copied_file_path)
                logging.debug(f"Cleaned up temporary copy: {copied_file_path}")
            except Exception as e:
                logging.error(f"Error cleaning up temporary copy {copied_file_path}: {e}")


    # Return the extracted header information
    # table_type is collected in table_features and returned here
    # (is_sheet, sheet_rows, row_contents, col_a_contents, last_bold_row, col_a_none_counts, col_contents, sheetnames) are features, not headers themselves
    # The original function returned table_type, all_sheet_row_headers, all_sheet_col_headers
    return table_type, all_sheet_row_headers, all_sheet_col_headers


# ---------------- Database Loading Function ------------------
def load_file_list_from_mysql_content() -> list:
    """
    Loads a list of table files (file_id, detect_format, full_path) from the database.
    Filters by supported formats (csv, xls, xlsx).
    """
    conn = None
    cursor = None
    results_with_path = []
    try:
        # Get database connection
        conn = pymysql.connect(**MYSQL_CONFIG)
        cursor = conn.cursor()

        # !! Modified SQL query to select table files !!
        # Select file_id, detect_format, and data_filename from the specified table
        # Filter by supported formats using the SUPPORTED_FORMATS constant
        formats_tuple = tuple(SUPPORTED_FORMATS)
        query = f"""
                SELECT file_id, detect_format, data_filename
                FROM ntcir_datafile
                WHERE detect_format IN %s
            """
        # Optional: LIMIT clause for testing
        # query += " LIMIT 100"

        cursor.execute(query, (formats_tuple,))
        # Fetch all matching rows
        results = cursor.fetchall()

        # Iterate through database results to validate file paths
        for file_id, detect_format, data_filename in results:
            # Basic validation: check if format is supported and filename exists
            if data_filename and detect_format in SUPPORTED_FORMATS:
                # Construct the full path to the file
                full_path = os.path.join(DATA_DIR, data_filename)
                # Check if the file exists on disk
                if not os.path.exists(full_path):
                    logging.warning(f"File not found on disk, skipping: {full_path} (ID: {file_id})")
                    continue # Skip this entry if the file doesn't exist

                # Add the valid file info to the results list
                results_with_path.append((file_id, detect_format, full_path))
            else:
                logging.warning(f"Skipping database entry (invalid format or missing filename): ID {file_id}, Format {detect_format}, Filename {data_filename}")

        logging.info(f"Loaded {len(results_with_path)} valid table file entries from MySQL.")
        return results_with_path

    except pymysql.Error as e:
        logging.error(f"Database error occurred: {e}")
        return [] # Return empty list on database error
    except Exception as e:
        logging.error(f"An unexpected error occurred while connecting to MySQL or querying: {e}")
        return [] # Return empty list on other errors
    finally:
        # Ensure cursor and connection are closed
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# ---------------- EDP Extraction Worker Function ------------------
def extract_edp_worker(args: Tuple[Any, str, str]) -> Tuple[Any, Optional[set]]:
    """
    Worker function to extract Entry Data Points (EDPs) from a single table file.
    Based on the file format, it calls the appropriate header extraction logic.
    EDPs are typically column or row headers.
    Returns a tuple (file_id, set of EDPs or None).
    Returns None if extraction fails.
    """
    file_id, detect_format, file_path = args
    combined_edp_set = set() # Initialize an empty set for extracted EDPs

    try:
        # Process based on the detected file format
        if detect_format == 'csv':
            try:
                # Read only the header row of the CSV file
                # Use low_memory=False to avoid potential DtypeWarning with mixed types
                # Try UTF-8 first, then Latin-1 as a fallback
                try:
                    header_df = pd.read_csv(file_path, header=0, nrows=0, encoding='utf-8', low_memory=False)
                except UnicodeDecodeError:
                    logging.debug(f"CSV {file_path} UTF-8 decode failed, trying latin-1.")
                    header_df = pd.read_csv(file_path, header=0, nrows=0, encoding='latin1', low_memory=False)

                # Get column names, convert to string, strip whitespace, and filter out empty names
                combined_edp_set = {str(col).strip() for col in header_df.columns if pd.notna(col) and str(col).strip()}
                # logging.debug(f"WORKER: Extracted {len(combined_edp_set)} EDPs from CSV header for ID {file_id}.")
                return file_id, combined_edp_set # Return file_id and the set of EDPs

            except Exception as e:
                 logging.error(f"WORKER: Error reading or processing CSV header for ID {file_id} ({file_path}): {e}")
                 return file_id, None # Return None on processing failure


        elif detect_format == 'xls' or detect_format == 'xlsx':
            # Use the extract_header function which handles XLS conversion and XLSX feature extraction
            table_type, all_sheet_row_headers, all_sheet_col_headers = extract_header(file_path, detect_format, TEMP_DIR)

            # Combine row and column headers from all sheets into a single set of EDPs
            if all_sheet_row_headers:
                for sheet_row_headers in all_sheet_row_headers:
                    if sheet_row_headers: # Check if the sheet had any extracted row headers
                        for row_headers in sheet_row_headers:
                             # Add headers from this row to the combined set
                            combined_edp_set.update(set(row_headers))

            if all_sheet_col_headers:
                for sheet_col_headers in all_sheet_col_headers:
                     if sheet_col_headers: # Check if the sheet had any extracted col headers
                         for col_headers in sheet_col_headers:
                             # Add headers from this column to the combined set
                             combined_edp_set.update(set(col_headers))

            # logging.debug(f"WORKER: Extracted {len(combined_edp_set)} EDPs from Excel headers for ID {file_id}.")
            return file_id, combined_edp_set # Return file_id and the set of EDPs

        else:
            # This case should ideally be prevented by filtering supported formats beforehand
            logging.warning(f"WORKER: Skipping file with unsupported format: ID {file_id}, Format {detect_format}")
            return file_id, None # Return None for unsupported formats

    except FileNotFoundError:
        logging.error(f"WORKER: File not found for ID {file_id}: {file_path}")
        return file_id, None
    except Exception as e:
        # Catch any unexpected error during the extraction process
        logging.error(f"WORKER: Unexpected error processing file ID: {file_id} ({file_path}): {e}", exc_info=False)
        return file_id, None # Return None on unexpected error


# ---------------- Main Execution Function ------------------
def run():
    """
    Main function to load table file list from the database,
    extract EDPs for each file, and save the results to a Pickle file.
    """
    logging.info("Starting table pattern extraction process.")

    # Dictionary to store results: {file_id: set(edps)}
    table_pattern_results = {}

    # Load the list of table files from the database
    all_files_info = load_file_list_from_mysql_content()

    if not all_files_info:
        logging.warning("No table files loaded from the database. Exiting.")
        return # Exit if no files to process

    logging.info(f"Processing {len(all_files_info)} table files.")

    # Iterate through each file info loaded from the database
    # Note: The original code processed these sequentially using the worker function directly.
    # If parallelism is desired, a ProcessPoolExecutor would be used here.
    # Keeping the original sequential processing logic for this refactoring.
    for files_info in all_files_info:
        file_id, combined_edp_set = extract_edp_worker(files_info)
        # Store the extracted EDPs in the results dictionary if extraction was successful
        if combined_edp_set is not None:
             table_pattern_results[file_id] = combined_edp_set
             logging.debug(f"Extracted EDPs for file ID {file_id}.")
        else:
             logging.warning(f"EDP extraction failed or returned empty for file ID {file_id}. Skipping result storage.")


    # Ensure the output directory exists before saving the file
    output_dir = os.path.dirname(OUTPUT_FILE)
    os.makedirs(output_dir, exist_ok=True)
    logging.info(f"Output directory '{output_dir}' exists or was created.")

    # Save the results dictionary to a Pickle file
    try:
        with open(OUTPUT_FILE, 'wb') as f:
            pickle.dump(table_pattern_results, f)
        logging.info(f"Table pattern results saved to {OUTPUT_FILE}. Total entries: {len(table_pattern_results)}")
    except Exception as e:
        logging.error(f"Error saving results to pickle file {OUTPUT_FILE}: {e}")


    logging.info("Table pattern extraction process finished.")


# ---------------- Main Entry Point ------------------
if __name__ == "__main__":
    # This is the main entry point when the script is executed directly.
    # It calls the 'run' function to start the process.
    
    run()